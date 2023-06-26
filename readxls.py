import pandas as pd
import re
import math
import os
from openpyxl import Workbook

def readUserId(df, fileName):
    columnNamesList = df.columns.tolist()
    try:
        pattern = re.compile(r'.*user\s*id.*', re.IGNORECASE)
        columnName = ""
        for key in columnNamesList:
            if pattern.search(key):
                print(key)
                columnName = key
                break

        if(columnName==""):
            print ("USER ID not found in ", fileName)
            return []
        useridList = df[key].values
        useridList = [x for x in useridList if type(x) is type("str")]

        print(useridList)
        return useridList

    except Exception as err:
        print(err)

def readFirstName(df, fileName):
    columnNamesList = df.columns.tolist()
    try:
        pattern = re.compile(r'.*first\s*name.*', re.IGNORECASE)
        columnName = ""
        for key in columnNamesList:
            if pattern.search(key):
                print(key)
                columnName = key
                break

        if(columnName==""):
            print ("First Name not found in ", fileName)
            return []
        fnList = df[key].values
        fnList = [x for x in fnList if type(x) is type("str")]
        print(fnList)
        return fnList

    except Exception as err:
        print(err)

def readLastName(df, fileName):
    columnNamesList = df.columns.tolist()
    try:
        pattern = re.compile(r'.*last\s*name.*', re.IGNORECASE)
        columnName = ""
        for key in columnNamesList:
            if pattern.search(key):
                print(key)
                columnName = key
                break

        if(columnName==""):
            print ("Last Name not found in ", fileName)
            return []
        lnList = df[key].values
        lnList = [x for x in lnList if type(x) is type("str")]
        print(lnList)
        return lnList

    except Exception as err:
        print(err)

def readFullName(df, fileName):
    columnNamesList = df.columns.tolist()
    try:
        pattern = re.compile(r'.*name.*', re.IGNORECASE)
        columnName = ""
        for key in columnNamesList:
            if pattern.search(key):
                print(key)
                columnName = key
                break

        if(columnName==""):
            print ("Full Name not found in ", fileName)
            return []
        lnList = df[key].values
        lnList = [x for x in lnList if type(x) is type("str")]
        print(lnList)
        return lnList

    except Exception as err:
        print(err)

def readDSId(df, fileName):
    columnNamesList = df.columns.tolist()
    try:
        pattern = re.compile(r'.*ds\s*id.*', re.IGNORECASE)
        columnName = ""
        for key in columnNamesList:
            if pattern.search(key):
                print(key)
                columnName = key
                break

        if(columnName==""):
            print ("DS ID not found in ", fileName)
            return []
        dsidList = df[key].values
        dsidList = [x for x in dsidList if type(x) is type("str")]
        print(dsidList)
        return dsidList


    except Exception as err:
        print(err)

def readCargillId(df, fileName):
    columnNamesList = df.columns.tolist()
    try:
        pattern = re.compile(r'.*car.', re.IGNORECASE)
        columnName = ""
        for key in columnNamesList:
            if pattern.search(key):
                print(key)
                columnName = key
                break

        if(columnName==""):
            print ("Cargill ID not found in ", fileName)
            return []
        dsidList = df[key].values
        dsidList = [x for x in dsidList if type(x) is type("str")]
        print(dsidList)
        return dsidList


    except Exception as err:
        print(err)

# def readBankUserId(df, fileName):
#     columnNamesList = df.columns.tolist()
#     try:
#         pattern = re.compile(r'.*bank\s*user\s*id.*', re.IGNORECASE)
#         columnName = ""
#         for key in columnNamesList:
#             if pattern.search(key):
#                 print(key)
#                 columnName = key
#                 break
#
#         if(columnName==""):
#             print ("Bank User ID not found in ", fileName)
#             return []
#         BUidList = df[key].values
#         BUidList = [x for x in BUidList if type(x) is type("str")]
#         print(BUidList)
#         return BUidList
#
#     except Exception as err:
#         print(err)

def readUserStatus(df, fileName):
    columnNamesList = df.columns.tolist()
    try:
        pattern = re.compile(r'.*status.*', re.IGNORECASE)
        columnName = ""
        for key in columnNamesList:
            if pattern.search(key):
                print(key)
                columnName = key
                break

        if(columnName==""):
            print ("User Status not found in ", fileName)
            return []
        statusList = df[key].values
        statusList = [x for x in statusList if type(x) is type("str")]
        print(statusList)
        return statusList

    except Exception as err:
        print(err)

def readEmail(df, fileName):
    columnNamesList = df.columns.tolist()
    emailidList = []
    try:
        pattern = re.compile(r'.*email\s*id.*', re.IGNORECASE)
        columnName = ""

        for key in columnNamesList:
            if pattern.search(key):
                print(key)
                columnName = key
                break
        if columnName=="":
            for col_name in df.columns:
                # if "manager" in col_name:
                    # continue
                col_data = df[col_name]

                for data in col_data:
                    if(type(data) == type("str")):
                        # print(data)
                        if "@cargill.com" in data or "@fisglobal.com" in data or "@crgl-thirdparty.com" in data or "@diamondv.com" in data or "@py.ey.com" in data:
                            emailidList.append(data)
                            # print(emailidList)
            print(emailidList)
            return emailidList


        if(columnName==""):
            print ("Email ID not found in ", fileName)
            return []
        emailidList = df[key].values
        emailidList = [x for x in emailidList if type(x) is type("str")]
        print(emailidList)
        return emailidList

    except Exception as err:
        print(err)

def readxls(filename):

    df = pd.read_excel(filename)

    columnNamesList = df.columns.tolist()
    sheetname = filename.split("_")[0]
    print(sheetname)
    writeOutputToFile(df, filename)

def traverseDir(dirPath):

# List to hold all Excel data frames
    all_data_frames = []

    # Loop through all files in the directory
    for fileName in os.listdir(dirPath):
        # Check if file is an Excel file
        if fileName.startswith("~$"):
            continue
        if fileName.endswith('.xlsx') or fileName.endswith('.xls'):
            # Read Excel file and append data frame to list
            filePath = os.path.join(dirPath, fileName)
            print(fileName)
            readxls(filePath)

def writeOutputToFile(df, filename):

    print("Columns List: ", df.columns.tolist())
    print()
    print()

    workbook = Workbook()
    sheet = workbook.active

    columnTitle = ("DS ID","BANK USER ID", "USER NAME", "USER EMAIL", "USER STATUS", "BANK NAME")

    # sheet.append(columnTitle)
    dsIdColumn = []
    dsIdColumn = readDSId(df, filename)
    if len(dsIdColumn) == 0:
        dsIdColumn = readCargillId(df, filename)
    
    userIdColumn = []
    userIdColumn = readUserId(df, filename)

    index = 0
    for id in userIdColumn:
        if not id[0].isdigit() and id[1].isdigit() or len(id) <= 9:
            if (index >= len(dsIdColumn)):
                dsIdColumn.append("")
            dsIdColumn[index] = id
            userIdColumn[index] = ""
        else:
            if (index >= len(dsIdColumn)):
                dsIdColumn.append("")
        index = index + 1

    emailIdColumn = readEmail(df, filename)
    first_names = readFirstName(df, filename)
    last_names = readLastName(df, filename)
    fullNameCol = readFullName(df, filename)

    userNameColumn = []
    if(len(fullNameCol)==0):
        if not first_names == 0 and not last_names == 0:
            userNameColumn = [first.strip() + " " + last.strip() for first, last in zip(first_names, last_names)]
        elif not last_names == 0:
            userNameColumn = first_names
        else:
            userNameColumn = last_names
    else:
        userNameColumn = fullNameCol
    # bankUserIdColumn = readBankUserId(df, filename)
    userStatus = readUserStatus(df, filename)

    print("PRINTING")
    print(dsIdColumn)
    print(userIdColumn)
    print(emailIdColumn)
    print(userNameColumn)
    print(userStatus)

    arrayOfColumns = [dsIdColumn, userIdColumn, userNameColumn, emailIdColumn, userStatus]

    largest_size = max(arrayOfColumns, key=len)
    size = len(largest_size)

    bankNameArr = []
    bankNameArr = [os.path.basename(filename)] * size
    arrayOfColumns.append(bankNameArr)


    print("SAVING DATA FOR " + filename)
    for col_idx, header in enumerate(columnTitle, start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = header

    for col_idx, column in enumerate(arrayOfColumns, start=1):
        if (not column):
            continue
        else:
            for row_idx, value in enumerate(column, start=2):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = value


    workbook.save('outputFolder/output_'+os.path.basename(filename))

if __name__ == '__main__':

    folder_name = "outputFolder"

    # Specify the path where you want to create the folder
    path = os.getcwd()

    # Create the full path by combining the parent folder path and the new folder name
    folder_path = os.path.join(path, folder_name)

    # Use the makedirs() function to create the folder
    os.makedirs(folder_path)


    traverseDir(os.getcwd())
